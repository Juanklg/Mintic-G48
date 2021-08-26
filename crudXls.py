import funciones
from datetime import datetime
from openpyxl import load_workbook

def Create(ruta:str, datos:dict):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['Datos del crud']
    # print('F'+str(Hoja_datos.max_row+1))
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row+1)]
    hoja=Archivo_Excel.active    

    task = [(2,"Nombre"),(3,"Descripcion"),(4,"Estado"),(5,"Inicio"),(6,"Fin")]
    for i in Hoja_datos:        
        if not(isinstance(i[0].value, int)):
            identificador  =i[0].row
            hoja.cell(row=identificador, column=1).value=identificador-1
            hoja.cell(row=identificador, column=task[0][0]).value=datos['Nombre']
            hoja.cell(row=identificador, column=task[1][0]).value=datos['Descripcion']
            hoja.cell(row=identificador, column=task[2][0]).value=datos['Estado']
            hoja.cell(row=identificador, column=task[3][0]).value=datos['Inicio']
            hoja.cell(row=identificador, column=task[4][0]).value=datos['Fin']
        break
    Archivo_Excel.save(ruta)
    return

def Read(ruta:str, filtro:str):
    task = [(2,"Nombre"),(3,"Descripcion"),(4,"Estado"),(5,"Inicio"),(6,"Fin")]
    print(ruta,filtro)
    Archivo_Excel = load_workbook(ruta)
    # print("Hojas dentro del archivo",Archivo_Excel.sheetnames)
    # wb.template = True
    # wb.save('document_template.xltx')
    Hoja_datos = Archivo_Excel['Datos del crud']
    # print("Hoja_datos.max_row",Hoja_datos.max_row)
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]
    info={}
    print("info ---------------- ", info)
    for i in Hoja_datos:
        if isinstance(i[0].value, int):
            info.setdefault(
                i[0].value,
                {
                    'Nombre':i[1].value, 
                    'Descripcion':i[2].value,
                    'Estado':i[3].value, 
                    'Inicio':i[4].value, 
                    'Fin':i[5].value
                })

    if not(filtro=='Todas'):
        info=filtrar(info,filtro)
    for i in info:
        print('****** Tarea ******')
        print(i,info[i])
        # detailVar(info)
        # print('Id:'+str(i)+'\n'+'titulo:'+str(info[i]['Tarea'])+'\n'+'Descripcion: '
        # +str(info)[i]['descripcion']) + '\n'+ 'Estado: '+str(info[i]['estado'])
        # +'\n'+'Fecha creacion: '+str(info[i]['Fecha de inicio'])
        # + '\n'+'Fecha de finalizacion: '+str(info[i]['Fecha de finalizacion'])
        print()

    return

def Update(ruta:str, id:int, taskDict:dict):
    task = [(2,"Nombre"),(3,"Descripcion"),(4,"Estado"),(5,"Inicio"),(6,"Fin")]
    Archivo_Exccel = load_workbook(ruta)
    Hoja_datos = Archivo_Exccel["Datos del crud"]
    Hoja_datos = Hoja_datos["A2":"F"+str(Hoja_datos.max_row)]
    hoja=Archivo_Exccel.active 
    
    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizado=6
    encontro=False 
    for i in Hoja_datos:
        if i[0].value==id:
            fila=i[0].row
            encontro=True
            for d in taskDict:                
                if d=="Nombre" and not(taskDict[d]==""):
                    hoja.cell(row=fila, column=titulo).value=taskDict[d]
                elif d=="Descripcion" and not (taskDict[d]==""):
                    hoja.cell(row=fila, column=descripcion).value=taskDict[d]
                elif d=="Estado" and not(taskDict[d]==""):
                    hoja.cell(row=fila, column=estado).value=taskDict[d]
                elif d=="Inicio" and not(taskDict[d]==""):
                    hoja.cell(row=fila, column=fecha_inicio).value=taskDict[d]
                elif d=="Fin" and not (taskDict[d]==""):
                    hoja.cell(row=fila, column=fecha_finalizado).value=taskDict[d]
    Archivo_Exccel.save(ruta)
    if encontro==False:
        print("Error : No existe una tarea con ese Id")
        print()
    return

def Delete(ruta, identificador):
    Archivo_Exccel= load_workbook(ruta)
    Hoja_datos = Archivo_Exccel["Datos del crud"]
    Hoja_datos=Hoja_datos["A2":"F"+str(Hoja_datos.max_row)]
    hoja=Archivo_Exccel.active
    task = [(2,"Nombre"),(3,"Descripcion"),(4,"Estado"),(5,"Inicio"),(6,"Fin")]
    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizado=6
    encontro=False    
    for i in Hoja_datos:        
        if i[0].value==identificador:
            fila=i[0].row
            print("i",i[0].row)
            encontro=True
            hoja.cell(row=fila,column=1).value=""
            hoja.cell(row=fila,column=task[0][0]).value=""
            hoja.cell(row=fila, column=task[1][0]).value=""
            hoja.cell(row=fila, column=task[2][0]).value = ""
            hoja.cell(row=fila, column=task[3][0]).value = ""
            hoja.cell(row=fila, column=task[4][0]).value = ""
    Archivo_Exccel.save(ruta)
    if encontro==False:
        print("Error: No Existe una tarea con ese id")
        print()
    return
